import openpyxl

def find_multiple_clicked_links(input_path, output_path):
    # Open the Excel spreadsheet
    wb = openpyxl.load_workbook(input_path)
    sheet = wb.active
    
    # Find the indices of columns A (email) and B (message)
    email_column = None
    message_column = None
    for col in range(1, sheet.max_column + 1):
        if sheet.cell(row=1, column=col).value == "email":
            email_column = col
        elif sheet.cell(row=1, column=col).value == "message":
            message_column = col
            
    # Check if columns are found
    if email_column is None or message_column is None:
        print("Columns not found. Make sure column names 'email' and 'message' are correct.")
        return
    
    # Dictionary to store email addresses with their associated "Clicked Link" count
    clicked_links_count = {}
    
    # Iterate through rows and count "Clicked Link" for each email address
    for row in range(2, sheet.max_row + 1):
        email = sheet.cell(row=row, column=email_column).value
        message = sheet.cell(row=row, column=message_column).value
        
        if message == "Clicked Link":
            if email in clicked_links_count:
                clicked_links_count[email] += 1
            else:
                clicked_links_count[email] = 1
    
    # Prompt the user to enter the output file path
    output_path = input("Enter the path for the output text file: ")

    # Write email addresses with more than one "Clicked Link" to a text file
    with open(output_path, "w") as output_file:
        for email, count in clicked_links_count.items():
            if count > 1:
                output_file.write(email + "\n")

    print("Output saved to", output_path)

# Example usage
input_path = input("Enter the path to the Excel spreadsheet: ")
find_multiple_clicked_links(input_path, "")
